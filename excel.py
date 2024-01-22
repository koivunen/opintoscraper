import openpyxl
import constants
import logging
log = logging.getLogger(__name__)
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path

import shutil

wb=False
applicants_file = Path(constants.OUTPUT_PATH) / 'applicants_automated.xlsx'
applicants_template = Path("applicants_template.xlsx")

assert applicants_template.exists()

FN=str(applicants_file)

if not applicants_file.exists():
    shutil.copy(applicants_template, applicants_file)
    log.warning("Generated %s",FN)

db_keys={}
def _build_key_index(sheet):
    for name in sheet['1']:
        db_keys[name.value]=name

def key_id(key):
    return get_column_letter(db_keys[key].column)
def key_index(key):
    return db_keys[key].column

def get_worksheet():
    global wb
    load=not wb
    if load:
        wb = load_workbook(filename = FN)
    sheet = wb["Sheet1"]
    if load:
        _build_key_index(sheet)
        _generate_app_cache()
    return sheet
def save():
    if wb:
        wb.save(FN)
        return True

applicants_cache={}
def _generate_app_cache():
    ws=get_worksheet()
    keyid=key_id("oid")
    for app in ws[keyid]:
        if app.value:
            applicants_cache[app.value]=app

#TODO: Precache (slow)
def find_applicant(oid):

    oid=oid.strip()
    for oid_fuzzy,app in applicants_cache.items():
        if oid in oid_fuzzy:
            return app.row

def cache_applicant(app,oid=None):
    val=oid or app.value
    if val:
        applicants_cache[val]=app
    else:
        log.error("Could not cache %s",str(oid))

def get_empty_row():
    ws=get_worksheet()
    for cell in ws["A"]:
        if not cell.value:
            return cell.row
    else:
        return cell.row + 1

import urllib.parse

def set_applicant(oid,preferred_name,last_name,path,target_oid_name):
    ws=get_worksheet()
    row=find_applicant(oid)
    new_applicant = not row
    if new_applicant:
        row = get_empty_row()
    oidcell = ws.cell(row=row,column=key_index("oid"))

    if new_applicant:
        # maybe modified accidentally, etc
        oidcell.value=oid
        cache_applicant(oidcell,oid)
        oidcell.style = "Hyperlink"

    oidcell.hyperlink = "https://virkailija.opintopolku.fi/cas/login?service=https://virkailija.opintopolku.fi/lomake-editori/applications/search?term="+urllib.parse.quote_plus(oid.strip())


    ws.cell(row=row,column=key_index("last-name")).value=last_name
    ws.cell(row=row,column=key_index("preferred-name")).value=preferred_name

    acell = ws.cell(row=row,column=key_index(target_oid_name))
    if not acell.value:
        acell.value='X'
        acell.style = "Hyperlink"
    acell.hyperlink = '\\\\utu.fi\\taltio\\SCITECH Masters admissions\\att\\kv_24\\' + path


if __name__ == "__main__":

    import coloredlogs

    coloredlogs.install(level='DEBUG')
    set_applicant("oid1337","lastname1","preferredname1","\\\\utu.fi\\taltio","suscity")
    save()
